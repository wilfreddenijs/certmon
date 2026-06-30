"""
CertMon - TLS Certificate Monitor & ACME Renewal Tool
Scans network devices for TLS certificates and manages renewals via ACME/Let's Encrypt
"""

import sys
import ssl
import socket
import json
import os
import subprocess
import threading
import ipaddress
import io
import base64
import uuid
from pathlib import Path
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, render_template, request, jsonify, send_file, Response

from certmon.config import resolve_data_dir
from certmon.ca_migration import migrate_legacy_ca_if_present
from certmon.db import Database
from certmon.artifacts import ArtifactStore
from certmon.acme_service import (
    ACMEAccountService,
    ACMEOrderService,
    NativeACMEAccountClient,
    NativeACMEOrderClient,
)
from certmon.deployment import DeploymentService, ExtronDeploymentAdapter
from certmon.dns.cloudflare import CloudflareDNSProvider, CloudflareError
from certmon.dns.manual import ManualDNSProvider
from certmon.external_ca import ExternalCAService
from certmon.local_ca import LocalCAService
from certmon.naming import safe_slug
from certmon.permissions import Permission, authorize
from certmon.renewals import ACMERenewalOrchestrator, RenewalService, StagingRequired
from certmon.toolbelt import ToolbeltBatchService
from certmon.vault import MemoryKeyProtector, Vault, WindowsDpapiProtector


def resource_path(relative):
    """Get absolute path — works for dev and PyInstaller bundle."""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative)


def data_dir():
    """Return CertMon's writable server data directory."""
    path = resolve_data_dir(
        frozen=getattr(sys, "frozen", False),
        executable=Path(sys.executable),
        source_dir=Path(__file__).resolve().parent,
    )
    path.mkdir(parents=True, exist_ok=True)
    return str(path)


app = Flask(__name__, template_folder=resource_path("templates"))

# Persistent storage
DATA_FILE = os.path.join(data_dir(), "certmon_data.json")
DB_FILE = os.path.join(data_dir(), "certmon.db")
database = Database(Path(DB_FILE))
database.initialize()
legacy_base = (
    Path(sys.executable).parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)
legacy_data_file = legacy_base / "certmon_data.json"
migration_source = None
for migration_source in (Path(DATA_FILE), legacy_data_file):
    if migration_source.exists():
        database.migrate_legacy_nonsecrets(migration_source)
        break
else:
    migration_source = None

def bootstrap_services(
    database,
    root,
    *,
    key_protector=None,
    deployment_adapters=None,
):
    """Build the secure services once, with injectable test/dev dependencies."""
    if key_protector is None:
        if sys.platform == "win32":
            key_protector = WindowsDpapiProtector()
        elif os.environ.get("CERTMON_DATA_DIR"):
            # Non-Windows test/dev runs use an isolated data directory. The memory
            # protector is deliberately never selected for the packaged app.
            key_protector = MemoryKeyProtector()
        else:
            return {}

    root = Path(root)
    vault = Vault(root / "secrets", key_protector)
    vault.initialize()
    artifact_store = ArtifactStore(root / "certificates", vault)
    migrate_legacy_ca_if_present(
        artifact_store,
        (Path(r"C:\CertMon\CA"), root / "CA"),
    )
    local_ca_service = LocalCAService(database, artifact_store)
    external_ca_service = ExternalCAService(database, artifact_store)
    acme_account_service = ACMEAccountService(
        database, vault, NativeACMEAccountClient
    )
    acme_order_service = ACMEOrderService(
        database, acme_account_service, NativeACMEOrderClient
    )
    dns_providers = {"manual": ManualDNSProvider()}
    try:
        dns_providers["cloudflare"] = CloudflareDNSProvider.load(database, vault)
    except CloudflareError:
        pass
    acme_orchestrator = ACMERenewalOrchestrator(
        database,
        renewal_service,
        acme_account_service,
        acme_order_service,
        artifact_store,
        dns_providers,
    )
    deployment_service = DeploymentService(
        database,
        artifact_store,
        renewal_service,
        adapters=(
            {"extron": ExtronDeploymentAdapter()}
            if deployment_adapters is None
            else deployment_adapters
        ),
    )
    toolbelt_service = ToolbeltBatchService(database, artifact_store, vault)
    return {
        "vault": vault,
        "artifact_store": artifact_store,
        "local_ca_service": local_ca_service,
        "external_ca_service": external_ca_service,
        "renewal_service": renewal_service,
        "acme_account_service": acme_account_service,
        "acme_order_service": acme_order_service,
        "acme_orchestrator": acme_orchestrator,
        "deployment_service": deployment_service,
        "toolbelt_service": toolbelt_service,
    }


vault = None
artifact_store = None
local_ca_service = None
external_ca_service = None
renewal_service = RenewalService(database)
acme_account_service = None
acme_order_service = None
acme_orchestrator = None
deployment_service = None
toolbelt_service = None
_services = bootstrap_services(database, data_dir())
globals().update(_services)
if migration_source is not None and vault is not None:
    database.complete_legacy_secret_migration(migration_source, vault)
ACME_STAGING = "https://acme-staging-v02.api.letsencrypt.org/directory"
ACME_PROD = "https://acme-v02.api.letsencrypt.org/directory"

COMMON_TLS_PORTS = [443, 8443, 8080, 4443, 9443, 4523]

scan_progress = {"running": False, "progress": 0, "total": 0, "current": ""}


def load_data():
    data = database.load_legacy_state()
    if vault is not None:
        for device in data.get("upload_devices", []):
            device_id = device.get("id") or f"legacy:{device.get('host')}"
            blob = database.get_secret(f"device-password:{device_id}")
            if blob is not None:
                device["password"] = vault.decrypt(
                    blob, purpose="device-password"
                ).decode("utf-8")
    return data


def save_data(data):
    sanitized = json.loads(json.dumps(data))
    for device in sanitized.get("upload_devices", []):
        password = device.pop("password", None)
        if password is None:
            continue
        if vault is None:
            raise RuntimeError("Secure credential storage is unavailable")
        device_id = device.get("id") or f"legacy:{device.get('host')}"
        database.put_secret(
            f"device-password:{device_id}",
            vault.encrypt(password.encode("utf-8"), purpose="device-password"),
            {"device_id": device_id},
        )
    database.save_legacy_state(sanitized)


def _safe_job(job):
    allowed = {
        "id",
        "endpoint_host",
        "endpoint_port",
        "issuer_type",
        "state",
        "version",
        "identifiers",
        "profile",
        "environment",
        "dns_provider",
        "certificate_id",
        "error_code",
        "error_message",
        "created_at",
        "updated_at",
        "artifact_name",
        "dns_records",
        "visible",
        "replaced",
        "metadata",
    }
    return {key: value for key, value in job.items() if key in allowed}


def _safe_job_with_dns(job):
    result = _safe_job(job)
    if job["state"] == "awaiting_dns":
        records = database.get_setting(f"acme-dns:{job['id']}", [])
        result["dns_records"] = [
            {"fqdn": record["fqdn"], "value": record["value"]}
            for record in records
        ]
    return result


def get_cert_info(host, port=443, timeout=3):
    try:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        conn = ctx.wrap_socket(
            socket.create_connection((host, port), timeout=timeout),
            server_hostname=host
        )
        cert_der = conn.getpeercert(binary_form=True)
        conn.close()

        from cryptography import x509
        from cryptography.hazmat.backends import default_backend
        cert = x509.load_der_x509_certificate(cert_der, default_backend())

        now = datetime.now(timezone.utc)
        not_after = cert.not_valid_after_utc
        not_before = cert.not_valid_before_utc
        days_remaining = (not_after - now).days

        try:
            cn = cert.subject.get_attributes_for_oid(x509.NameOID.COMMON_NAME)[0].value
        except Exception:
            cn = host

        try:
            san_ext = cert.extensions.get_extension_for_class(x509.SubjectAlternativeName)
            sans = san_ext.value.get_values_for_type(x509.DNSName)
        except Exception:
            sans = []

        try:
            issuer = cert.issuer.get_attributes_for_oid(x509.NameOID.COMMON_NAME)[0].value
        except Exception:
            issuer = "Unknown"

        try:
            issuer_org = cert.issuer.get_attributes_for_oid(x509.NameOID.ORGANIZATION_NAME)[0].value
        except Exception:
            issuer_org = None

        # Self-signed: subject DN == issuer DN
        self_signed = cert.subject == cert.issuer
        cert_type = "Self-signed" if self_signed else f"CA: {issuer_org or issuer}"

        if days_remaining < 0:
            status = "expired"
        elif days_remaining <= 20:
            status = "critical"
        elif days_remaining <= 47:
            status = "warning"
        else:
            status = "ok"

        return {
            "host": host, "port": port, "cn": cn, "issuer": issuer,
            "issuer_org": issuer_org, "self_signed": self_signed,
            "cert_type": cert_type, "sans": sans,
            "not_before": not_before.isoformat(), "not_after": not_after.isoformat(),
            "days_remaining": days_remaining, "status": status,
            "serial": str(cert.serial_number),
            "last_checked": now.isoformat(), "error": None
        }
    except (socket.timeout, ConnectionRefusedError, OSError):
        return None
    except Exception as e:
        return {
            "host": host, "port": port, "cn": host, "issuer": "N/A", "sans": [],
            "not_before": None, "not_after": None, "days_remaining": None,
            "status": "error", "serial": None,
            "last_checked": datetime.now(timezone.utc).isoformat(), "error": str(e)
        }


def scan_host(host):
    results = []
    for port in COMMON_TLS_PORTS:
        info = get_cert_info(host, port)
        if info and info.get("error") is None:
            results.append(info)
    return results


def scan_range_worker(ip_range):
    global scan_progress
    try:
        network = ipaddress.ip_network(ip_range, strict=False)
        hosts = list(network.hosts())
        scan_progress["total"] = len(hosts)
        scan_progress["progress"] = 0
        found = []
        with ThreadPoolExecutor(max_workers=50) as executor:
            futures = {executor.submit(scan_host, str(ip)): str(ip) for ip in hosts}
            for future in as_completed(futures):
                ip = futures[future]
                scan_progress["current"] = ip
                scan_progress["progress"] += 1
                try:
                    results = future.result()
                    found.extend(results)
                except Exception:
                    pass
        return found
    except Exception:
        return []


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/data")
def api_data():
    return jsonify(load_data())


@app.route("/api/hosts", methods=["POST"])
def add_host():
    data = load_data()
    body = request.json
    host = body.get("host", "").strip()
    port = int(body.get("port", 443))
    if not host:
        return jsonify({"error": "No host provided"}), 400
    entry = {"host": host, "port": port}
    if entry not in data["manual_hosts"]:
        data["manual_hosts"].append(entry)
    info = get_cert_info(host, port)
    if info:
        data["certificates"][f"{host}:{port}"] = info
    save_data(data)
    return jsonify({"ok": True, "cert": info})


@app.route("/api/hosts/<path:host_port>", methods=["DELETE"])
def remove_host(host_port):
    data = load_data()
    parts = host_port.rsplit(":", 1)
    host = parts[0]
    port = int(parts[1]) if len(parts) > 1 else 443
    data["manual_hosts"] = [h for h in data["manual_hosts"]
                             if not (h["host"] == host and h["port"] == port)]
    data["certificates"].pop(f"{host}:{port}", None)
    save_data(data)
    return jsonify({"ok": True})


@app.route("/api/refresh/<path:host_port>", methods=["POST"])
def refresh_host(host_port):
    data = load_data()
    parts = host_port.rsplit(":", 1)
    host = parts[0]
    port = int(parts[1]) if len(parts) > 1 else 443
    info = get_cert_info(host, port)
    if info:
        data["certificates"][f"{host}:{port}"] = info
        save_data(data)
    return jsonify({"ok": True, "cert": info})


@app.route("/api/scan/ranges", methods=["POST"])
def add_range():
    data = load_data()
    body = request.json
    ip_range = body.get("range", "").strip()
    if not ip_range:
        return jsonify({"error": "No range provided"}), 400
    try:
        ipaddress.ip_network(ip_range, strict=False)
    except ValueError:
        return jsonify({"error": "Invalid CIDR range"}), 400
    if ip_range not in data["scan_ranges"]:
        data["scan_ranges"].append(ip_range)
        save_data(data)
    return jsonify({"ok": True})


@app.route("/api/scan/ranges/<path:ip_range>", methods=["DELETE"])
def remove_range(ip_range):
    data = load_data()
    data["scan_ranges"] = [r for r in data["scan_ranges"] if r != ip_range]
    save_data(data)
    return jsonify({"ok": True})


@app.route("/api/scan/start", methods=["POST"])
def start_scan():
    global scan_progress
    if scan_progress["running"]:
        return jsonify({"error": "Scan already running"}), 400
    data = load_data()
    ranges = data.get("scan_ranges", [])
    if not ranges:
        return jsonify({"error": "No scan ranges configured"}), 400

    def run_scan():
        global scan_progress
        scan_progress["running"] = True
        d = load_data()
        for ip_range in ranges:
            scan_progress["current"] = f"Scanning {ip_range}..."
            results = scan_range_worker(ip_range)
            for cert in results:
                key = f"{cert['host']}:{cert['port']}"
                d["certificates"][key] = cert
                entry = {"host": cert["host"], "port": cert["port"]}
                if entry not in d["manual_hosts"]:
                    d["manual_hosts"].append(entry)
            save_data(d)
        scan_progress["running"] = False
        scan_progress["current"] = "Done"

    threading.Thread(target=run_scan, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/scan/progress")
def scan_progress_api():
    return jsonify(scan_progress)


@app.route("/api/renew", methods=["POST"])
def create_renewal():
    authorize(Permission.ISSUE_CERTIFICATE)
    body = request.get_json(silent=True) or {}
    try:
        job = renewal_service.create_job(
            endpoint_host=body.get("endpoint_host") or body.get("host") or "",
            endpoint_port=body.get("endpoint_port", body.get("port", 443)),
            issuer_type=body.get("issuer_type") or "external_ca",
            identifiers=body.get("identifiers") or [body.get("host")],
            profile=body.get("profile") or "generic-rsa",
            environment=body.get("environment"),
            dns_provider=body.get("dns_provider"),
            metadata=body.get("metadata") if isinstance(body.get("metadata"), dict) else None,
        )
        return jsonify(_safe_job(job)), 201
    except (TypeError, ValueError) as error:
        return jsonify({"error": str(error)}), 400


@app.route("/api/renewals")
def list_renewals():
    return jsonify([_safe_job_with_dns(job) for job in database.list_jobs()])


@app.route("/api/renewals/<job_id>", methods=["GET"])
def renewal_detail(job_id):
    job = database.get_job(job_id)
    if job is None:
        return jsonify({"error": "Not found"}), 404
    result = _safe_job_with_dns(job)
    result["events"] = database.list_events(job_id)
    return jsonify(result)


@app.route("/api/renewals/<job_id>/start", methods=["POST"])
def start_renewal(job_id):
    authorize(Permission.ISSUE_CERTIFICATE)
    job = database.get_job(job_id)
    if job is None:
        return jsonify({"error": "Not found"}), 404
    body = request.get_json(silent=True) or {}
    try:
        if job["issuer_type"] == "acme":
            if acme_orchestrator is None:
                return jsonify({"error": "ACME service is unavailable"}), 503
            result = acme_orchestrator.start_acme(
                job_id,
                email=body.get("email", ""),
                terms_of_service_agreed=body.get("terms_of_service_agreed") is True,
            )
        elif job["issuer_type"] == "external_ca":
            if job.get("metadata", {}).get("external_ca_workflow") == "existing":
                return (
                    jsonify(
                        {
                            "error": "This draft is for importing an existing certificate and private key. Use Import certificate or delete the entry."
                        }
                    ),
                    400,
                )
            artifact_name = external_ca_service.create_csr_job(job_id)
            result = {**database.get_job(job_id), "artifact_name": artifact_name}
        elif job["issuer_type"] == "local_ca":
            issuing = renewal_service.transition(
                job_id, "draft", job["version"], "issuing"
            )
            issued = local_ca_service.issue(
                identifiers=tuple(job["identifiers"]),
                profile_name=job["profile"],
                device_name=job["endpoint_host"],
            )
            result = renewal_service.transition(
                job_id,
                "issuing",
                issuing["version"],
                "issued",
                {"certificate_id": issued["certificate_id"]},
            )
        else:
            raise ValueError("Unknown issuer type")
        return jsonify(_safe_job(result))
    except StagingRequired as error:
        return jsonify({"error": error.code, "action_url": error.action_url}), 409
    except (KeyError, TypeError, ValueError) as error:
        return jsonify({"error": str(error)}), 400


@app.route("/api/renewals/<job_id>/manual-dns/continue", methods=["POST"])
def continue_manual_dns(job_id):
    authorize(Permission.ISSUE_CERTIFICATE)
    if acme_orchestrator is None:
        return jsonify({"error": "ACME service is unavailable"}), 503
    try:
        return jsonify(_safe_job(acme_orchestrator.continue_manual_dns(job_id)))
    except (KeyError, ValueError) as error:
        return jsonify({"error": str(error)}), 400
    except Exception as error:
        return jsonify({"error": f"Could not continue renewal: {error}"}), 500


@app.route("/api/renewals/<job_id>/cancel", methods=["POST"])
def cancel_renewal(job_id):
    authorize(Permission.ISSUE_CERTIFICATE)
    try:
        job = database.get_job(job_id)
        service = acme_orchestrator if job and job["issuer_type"] == "acme" else renewal_service
        return jsonify(_safe_job(service.cancel(job_id)))
    except (KeyError, ValueError) as error:
        return jsonify({"error": str(error)}), 400


@app.route("/api/renewals/<job_id>", methods=["DELETE"])
def delete_renewal(job_id):
    authorize(Permission.ISSUE_CERTIFICATE)
    try:
        renewal_service.delete_terminal_job(job_id)
        return "", 204
    except KeyError:
        return jsonify({"error": "Not found"}), 404
    except ValueError as error:
        return jsonify({"error": str(error)}), 400


@app.route("/api/renewals/<job_id>/retry-cleanup", methods=["POST"])
def retry_renewal_cleanup(job_id):
    authorize(Permission.ISSUE_CERTIFICATE)
    try:
        return jsonify(_safe_job(acme_orchestrator.retry_cleanup(job_id)))
    except (KeyError, ValueError) as error:
        return jsonify({"error": str(error)}), 400


@app.route("/api/export/excel")
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = load_data()
    certs = list(data.get("certificates", {}).values())

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificates"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"CertMon — TLS Certificate Report  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(name="Arial", bold=True, size=13, color="00E5FF")
    title_cell.fill = PatternFill("solid", fgColor="0D1220")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    counts = {s: sum(1 for c in certs if c.get("status") == s)
              for s in ("ok", "warning", "critical", "expired", "error")}
    ws.merge_cells("A2:I2")
    ws["A2"].value = (f"Total: {len(certs)}   OK: {counts['ok']}   "
                      f"Warning: {counts['warning']}   Critical: {counts['critical']}   "
                      f"Expired: {counts['expired']}")
    ws["A2"].font = Font(name="Arial", size=10, color="444444")
    ws["A2"].fill = PatternFill("solid", fgColor="EEF2FF")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = ["Host", "Port", "Common Name", "Issuer", "Type", "Status",
               "Days Left", "Issued", "Expires", "Last Checked"]
    ws.row_dimensions[3].height = 22
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D3A5A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    order = {"critical": 0, "expired": 1, "warning": 2, "error": 3, "ok": 4}
    certs.sort(key=lambda c: order.get(c.get("status", "error"), 5))

    status_colors = {
        "ok":       ("C8F7DC", "1A7A3D"),
        "warning":  ("FFF3CC", "7A5200"),
        "critical": ("FFD6D6", "7A0000"),
        "expired":  ("EEEEEE", "555555"),
        "error":    ("F5F5F5", "888888"),
    }

    def fmt_date(iso):
        if not iso:
            return ""
        try:
            return datetime.fromisoformat(iso).strftime("%Y-%m-%d")
        except Exception:
            return iso

    for i, cert in enumerate(certs):
        row = i + 4
        status = cert.get("status", "error")
        bg, fg = status_colors.get(status, ("F5F5F5", "888888"))
        alt_bg = "FFFFFF" if i % 2 == 0 else "F8FAFC"
        days = cert.get("days_remaining")
        values = [cert.get("host", ""), cert.get("port", ""), cert.get("cn", ""),
                  cert.get("issuer", ""), cert.get("cert_type", "Unknown"), status.upper(),
                  str(days) if days is not None else "N/A",
                  fmt_date(cert.get("not_before")), fmt_date(cert.get("not_after")),
                  fmt_date(cert.get("last_checked"))]
        ws.row_dimensions[row].height = 18
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name="Arial", size=10,
                             color=fg if col == 5 else "222222", bold=(col == 5))
            cell.fill = PatternFill("solid", fgColor=bg if col == 5 else alt_bg)
            cell.alignment = Alignment(
                horizontal="center" if col in (2, 5, 6, 7, 8, 9) else "left",
                vertical="center")
            cell.border = border

    col_widths = [20, 8, 28, 30, 18, 12, 10, 14, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J{3 + len(certs)}"

    ws2 = wb.create_sheet("Renewals")
    r_headers = ["Host", "Port", "Method", "Environment", "Status", "Created", "Command"]
    for col, h in enumerate(r_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D3A5A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for i, r in enumerate(data.get("renewals", []), 2):
        vals = [r.get("host"), r.get("port"), r.get("method"),
                "Staging" if r.get("staging") else "Production",
                r.get("status", "").upper(), fmt_date(r.get("created")), r.get("command", "")]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=(col == 7))
            cell.border = border
    for i, w in enumerate([20, 8, 12, 14, 12, 14, 60], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"certmon_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from flask import Response
    response = Response(
        buf.getvalue(),
        status=200,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


# ─────────────────────────────────────────────
# CA Management
# ─────────────────────────────────────────────

def ca_exists():
    return bool(
        artifact_store
        and artifact_store.has_certificate(LocalCAService.CA_CERTIFICATE_ID)
    )


def _certificate_selector(metadata):
    identifiers = metadata.get("identifiers") or []
    for value in identifiers:
        try:
            ipaddress.ip_address(value)
        except ValueError:
            continue
        return value
    return identifiers[0] if identifiers else metadata.get("id")


def _filename_part(value):
    return safe_slug(value, max_length=80)


def _certificate_download_filename(certificate_id, artifact_name):
    metadata = database.get_certificate(certificate_id) if database else None
    metadata = metadata or {"id": certificate_id}
    identifiers = metadata.get("identifiers") or []
    parts = []
    for value in (metadata.get("device_name"), identifiers[0] if identifiers else None):
        part = _filename_part(value)
        if part and part not in parts:
            parts.append(part)
    profile = metadata.get("profile")
    if profile == "extron-rsa":
        parts.append("extron")
    elif profile:
        parts.append(_filename_part(profile))
    base = "-".join(part for part in parts if part) or "certificate"
    short_id = _filename_part(certificate_id)[:8]
    if short_id and short_id not in base:
        base = f"{base}-{short_id}"
    artifact_labels = {
        "certificate.pem": "certificate.pem",
        "chain.pem": "chain.pem",
        "full-chain.pem": "full-chain.pem",
        "request.csr": "request.csr",
        "private-key.pem": "private-key.pem",
        "combined.pem": "extron-combined.pem" if profile == "extron-rsa" else "combined.pem",
    }
    return f"{base}-{artifact_labels.get(artifact_name, artifact_name)}"


@app.route("/api/ca/status")
def ca_status():
    if not ca_exists():
        return jsonify({"exists": False})
    try:
        from cryptography import x509
        from cryptography.x509.oid import NameOID

        cert = x509.load_pem_x509_certificate(
            artifact_store.read_public(
                LocalCAService.CA_CERTIFICATE_ID, "certificate.pem"
            )
        )
        now = datetime.now(timezone.utc)
        days = (cert.not_valid_after_utc - now).days
        cn = cert.subject.get_attributes_for_oid(NameOID.COMMON_NAME)[0].value
        return jsonify({
            "exists": True,
            "cn": cn,
            "not_after": cert.not_valid_after_utc.isoformat(),
            "days_remaining": days,
            "certificate_id": LocalCAService.CA_CERTIFICATE_ID,
        })
    except Exception as e:
        return jsonify({"exists": False, "error": str(e)})


@app.route("/api/ca/generate", methods=["POST"])
def ca_generate():
    if local_ca_service is None:
        return jsonify({"error": "Secure certificate storage is unavailable"}), 503
    if ca_exists():
        return jsonify({"error": "CA already exists. Delete existing CA files first."}), 400
    try:
        return jsonify({"ok": True, **local_ca_service.generate_ca()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ca/install", methods=["POST"])
def ca_install():
    """Install CA cert into Windows trust store."""
    if not ca_exists():
        return jsonify({"error": "No CA found. Generate one first."}), 400
    if sys.platform != "win32":
        return jsonify({
            "ok": True,
            "method": "manual",
            "message": "Download the CA certificate and install it in your OS trust store.",
        })
    try:
        certificate_path = (
            artifact_store.root
            / LocalCAService.CA_CERTIFICATE_ID
            / "certificate.pem"
        )
        result = subprocess.run(
            ["certutil", "-addstore", "-user", "Root", str(certificate_path)],
            capture_output=True,
            text=True,
        )
        if result.returncode == 0:
            return jsonify({"ok": True, "method": "certutil", "output": result.stdout})
        return jsonify({"error": result.stderr or result.stdout}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ca/download-cert")
def ca_download_cert():
    """Download the CA certificate for manual installation."""
    if not ca_exists():
        return jsonify({"error": "No CA found"}), 404
    data = artifact_store.read_public(
        LocalCAService.CA_CERTIFICATE_ID, "certificate.pem"
    )
    response = Response(data, status=200, mimetype="application/x-pem-file")
    response.headers["Content-Disposition"] = 'attachment; filename="certmon-ca.crt"'
    return response


@app.route("/api/ca/issue", methods=["POST"])
def ca_issue():
    """Issue a device certificate signed by the local CA."""
    if not ca_exists():
        return jsonify({"error": "No CA found. Generate one first."}), 400
    if local_ca_service is None:
        return jsonify({"error": "Secure certificate storage is unavailable"}), 503
    try:
        body = request.get_json(silent=True) or {}
        ip = (body.get("ip") or "").strip()
        hostname = (body.get("hostname") or "").strip()
        identifiers = tuple(value for value in (hostname, ip) if value)
        if not identifiers:
            return jsonify({"error": "Provide at least an IP or hostname"}), 400
        result = local_ca_service.issue(
            identifiers=identifiers,
            profile_name=body.get("profile") or "extron-rsa",
            device_name=body.get("name") or hostname or ip,
        )
        return jsonify({"ok": True, "cn": hostname or ip, **result})
    except (KeyError, ValueError) as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ca/devices-txt")
def ca_devices_txt():
    """List of `ip,pemfile` for every issued device cert that has a .pem —
    ready to feed to toolbelt_uploader.py --list devices.txt.

    Two columns so each device maps to its EXACT .pem regardless of whether the
    cert was named by IP or hostname. Column 1 (the IP/host) is what Toolbelt
    selects by; column 2 is the cert to upload. One line per issued cert."""
    lines = []
    for metadata in database.list_certificates():
        if metadata.get("kind") != "leaf" or metadata.get("issuer_type") != "local_ca":
            continue
        selector = _certificate_selector(metadata)
        if selector:
            lines.append(f"{selector},{metadata['id']}")
    body = "\n".join(lines) + ("\n" if lines else "")
    response = Response(body, status=200, mimetype="text/plain")
    response.headers["Content-Disposition"] = 'attachment; filename="devices.txt"'
    return response


@app.route("/api/ca/issued/<certificate_id>", methods=["DELETE"])
def ca_delete_issued(certificate_id):
    """Delete an issued device cert and its companion files (.crt/.key/.pem,
    plus an _encrypted.key if present)."""
    metadata = database.get_certificate(certificate_id)
    if (
        metadata is None
        or metadata.get("kind") != "leaf"
        or metadata.get("issuer_type") != "local_ca"
    ):
        return jsonify({"error": "Not found"}), 404
    if artifact_store is not None:
        artifact_store.delete_certificate_set(certificate_id)
    database.delete_certificate(certificate_id)
    return jsonify({"ok": True, "removed": [certificate_id]})


@app.route("/api/ca/issued")
def ca_issued():
    """List issued device certs."""
    if artifact_store is None:
        return jsonify([])
    now = datetime.now(timezone.utc)
    certificates = []
    for metadata in database.list_certificates():
        if metadata.get("kind") != "leaf" or metadata.get("issuer_type") != "local_ca":
            continue
        not_after = datetime.fromisoformat(metadata["not_after"])
        certificates.append({
            "certificate_id": metadata["id"],
            "cn": metadata.get("identifiers", [metadata["id"]])[0],
            "not_after": metadata["not_after"],
            "days_remaining": (not_after - now).days,
            "profile": metadata.get("profile"),
        })
    return jsonify(certificates)


@app.route("/api/ca/download/<certificate_id>")
def ca_download_file(certificate_id):
    """Download a public device certificate. Private artifacts stay server-side."""
    if artifact_store is None or not artifact_store.has_certificate(certificate_id):
        return jsonify({"error": "Not found"}), 404
    try:
        data = artifact_store.read_public(certificate_id, "certificate.pem")
    except (FileNotFoundError, ValueError):
        return jsonify({"error": "Not found"}), 404
    response = Response(data, status=200, mimetype="application/x-pem-file")
    filename = _certificate_download_filename(certificate_id, "certificate.pem")
    filename = filename[:-4] + ".crt" if filename.endswith(".pem") else filename
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@app.route("/api/external-ca/trust-anchors", methods=["POST"])
def external_ca_import_trust_anchor():
    authorize(Permission.ISSUE_CERTIFICATE)
    if external_ca_service is None:
        return jsonify({"error": "Secure certificate storage is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    try:
        trust_anchor_id = external_ca_service.import_trust_anchor(
            body.get("trust_anchor_id", "").strip(),
            body.get("certificate_pem", "").encode("utf-8"),
        )
        return jsonify({"ok": True, "trust_anchor_id": trust_anchor_id})
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/renewals/<job_id>/external/csr", methods=["POST"])
def external_ca_create_csr(job_id):
    authorize(Permission.ISSUE_CERTIFICATE)
    if external_ca_service is None:
        return jsonify({"error": "Secure certificate storage is unavailable"}), 503
    try:
        artifact_name = external_ca_service.create_csr_job(job_id)
        return jsonify({"ok": True, "artifact_name": artifact_name})
    except (KeyError, ValueError) as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/renewals/<job_id>/external/csr", methods=["GET"])
def external_ca_download_csr(job_id):
    authorize(Permission.DOWNLOAD_PUBLIC_CERTIFICATE)
    if external_ca_service is None:
        return jsonify({"error": "Secure certificate storage is unavailable"}), 503
    try:
        data = external_ca_service.read_csr(job_id)
    except (FileNotFoundError, KeyError, ValueError):
        return jsonify({"error": "CSR not found"}), 404
    response = Response(data, status=200, mimetype="application/pkcs10")
    response.headers["Content-Disposition"] = f'attachment; filename="{job_id}.csr"'
    return response


@app.route("/api/renewals/<job_id>/external/complete", methods=["POST"])
def external_ca_complete(job_id):
    authorize(Permission.ISSUE_CERTIFICATE)
    if external_ca_service is None:
        return jsonify({"error": "Secure certificate storage is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    try:
        certificate_id = external_ca_service.complete_csr_job(
            job_id,
            body.get("certificate_pem", "").encode("utf-8"),
            (body.get("chain_pem") or "").encode("utf-8") or None,
            body.get("trust_anchor_id"),
        )
        return jsonify({"ok": True, "certificate_id": certificate_id})
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/renewals/<job_id>/external/import", methods=["POST"])
def external_ca_import_existing(job_id):
    authorize(Permission.ISSUE_CERTIFICATE)
    if external_ca_service is None:
        return jsonify({"error": "Secure certificate storage is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    try:
        certificate_id = external_ca_service.import_existing(
            job_id,
            body.get("certificate_pem", "").encode("utf-8"),
            (body.get("chain_pem") or "").encode("utf-8") or None,
            body.get("private_key_pem", "").encode("utf-8"),
            body.get("passphrase"),
            body.get("trust_anchor_id"),
        )
        return jsonify({"ok": True, "certificate_id": certificate_id})
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/credentials/cloudflare", methods=["GET"])
def get_cloudflare_credentials():
    authorize(Permission.MANAGE_DNS_CREDENTIALS)
    config = database.get_setting(CloudflareDNSProvider.SETTING_ID)
    return jsonify(
        {
            "configured": config is not None,
            "zones": config.get("zones", []) if config else [],
        }
    )


@app.route("/api/credentials/cloudflare", methods=["POST"])
def create_cloudflare_credentials():
    authorize(Permission.MANAGE_DNS_CREDENTIALS)
    if vault is None:
        return jsonify({"error": "Secure credential storage is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    try:
        CloudflareDNSProvider.configure(
            database,
            vault,
            token=body.get("token", ""),
            zones=body.get("zones") or [],
        )
        provider = CloudflareDNSProvider.load(database, vault)
        if acme_orchestrator is not None:
            acme_orchestrator.dns_providers["cloudflare"] = provider
        config = database.get_setting(CloudflareDNSProvider.SETTING_ID)
        return jsonify({"configured": True, "zones": config["zones"]}), 201
    except (TypeError, ValueError, CloudflareError):
        return jsonify({"error": "Cloudflare credentials are invalid"}), 400


@app.route("/api/credentials/cloudflare", methods=["DELETE"])
def delete_cloudflare_credentials():
    authorize(Permission.MANAGE_DNS_CREDENTIALS)
    database.delete_secret(CloudflareDNSProvider.SECRET_ID)
    database.delete_setting(CloudflareDNSProvider.SETTING_ID)
    if acme_orchestrator is not None:
        acme_orchestrator.dns_providers.pop("cloudflare", None)
    return "", 204


@app.route("/api/certificates/<certificate_id>/public/<artifact_name>")
def download_public_artifact(certificate_id, artifact_name):
    authorize(Permission.DOWNLOAD_PUBLIC_CERTIFICATE)
    allowed = {"certificate.pem", "chain.pem", "full-chain.pem", "request.csr"}
    if artifact_name not in allowed or artifact_store is None:
        return jsonify({"error": "Not found"}), 404
    try:
        data = artifact_store.read_public(certificate_id, artifact_name)
    except (FileNotFoundError, ValueError, PermissionError):
        return jsonify({"error": "Not found"}), 404
    response = Response(data, status=200, mimetype="application/x-pem-file")
    filename = _certificate_download_filename(certificate_id, artifact_name)
    response.headers["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )
    return response


@app.route("/api/certificates")
def list_deployable_certificates():
    authorize(Permission.DEPLOY_CERTIFICATE)
    certificates = []
    for metadata in database.list_certificates():
        if metadata.get("kind") != "leaf":
            continue
        certificates.append(
            {
                "certificate_id": metadata["id"],
                "identifiers": metadata.get("identifiers", []),
                "issuer_type": metadata.get("issuer_type"),
                "profile": metadata.get("profile"),
                "not_after": metadata.get("not_after"),
            }
        )
    return jsonify(certificates)


@app.route("/api/certificates/<certificate_id>/private/<artifact_name>")
def download_private_artifact(certificate_id, artifact_name):
    authorize(Permission.DOWNLOAD_PRIVATE_KEY)
    if artifact_name not in {"private-key.pem", "combined.pem"} or artifact_store is None:
        return jsonify({"error": "Not found"}), 404
    try:
        with artifact_store.materialize_private(certificate_id, artifact_name) as path:
            data = path.read_bytes()
    except (FileNotFoundError, ValueError, PermissionError):
        return jsonify({"error": "Not found"}), 404
    database.record_event(
        "private_artifact_downloaded",
        {"certificate_id": certificate_id, "artifact_name": artifact_name},
    )
    response = Response(data, status=200, mimetype="application/x-pem-file")
    filename = _certificate_download_filename(certificate_id, artifact_name)
    response.headers["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )
    return response


# ── Certificate Upload Module ─────────────────────────────────────────────────

# Toolbelt batch upload ------------------------------------------------------

def _toolbelt_unavailable():
    return toolbelt_service is None or artifact_store is None or vault is None


@app.route("/api/toolbelt/devices", methods=["GET"])
def toolbelt_devices():
    authorize(Permission.DEPLOY_CERTIFICATE)
    if _toolbelt_unavailable():
        return jsonify({"error": "Toolbelt batch upload is unavailable"}), 503
    return jsonify({"devices": toolbelt_service.list_devices()})


@app.route("/api/toolbelt/selection", methods=["PATCH"])
def toolbelt_selection():
    authorize(Permission.DEPLOY_CERTIFICATE)
    if _toolbelt_unavailable():
        return jsonify({"error": "Toolbelt batch upload is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    selectors = body.get("selectors")
    if selectors is None or not isinstance(selectors, list):
        return jsonify({"error": "selectors must be a list"}), 400
    toolbelt_service.save_selection([str(value) for value in selectors])
    return jsonify({"ok": True, "devices": toolbelt_service.list_devices()})


@app.route("/api/toolbelt/devices/<path:selector>/credentials", methods=["PATCH"])
def toolbelt_credentials(selector):
    authorize(Permission.DEPLOY_CERTIFICATE)
    if _toolbelt_unavailable():
        return jsonify({"error": "Toolbelt batch upload is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    username = (body.get("username") or "admin").strip()
    password = body.get("password")
    if password is None:
        return jsonify({"error": "password is required"}), 400
    toolbelt_service.save_credentials(
        selector, username=username, password=str(password)
    )
    return jsonify({"ok": True})


@app.route("/api/toolbelt/dry-run", methods=["POST"])
def toolbelt_dry_run():
    authorize(Permission.DEPLOY_CERTIFICATE)
    if _toolbelt_unavailable():
        return jsonify({"error": "Toolbelt batch upload is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    if any(key in body for key in ("private_key_pem", "combined_pem", "pem")):
        return jsonify({"error": "Private certificate material must stay server-side"}), 400
    try:
        selectors = body["selectors"] if "selectors" in body else None
        run = toolbelt_service.start(
            mode="dry-run", selectors=selectors
        )
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    return jsonify({"ok": True, "run": run})


@app.route("/api/toolbelt/upload", methods=["POST"])
def toolbelt_upload():
    authorize(Permission.DEPLOY_CERTIFICATE)
    if _toolbelt_unavailable():
        return jsonify({"error": "Toolbelt batch upload is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    if any(key in body for key in ("private_key_pem", "combined_pem", "pem")):
        return jsonify({"error": "Private certificate material must stay server-side"}), 400
    try:
        selectors = body["selectors"] if "selectors" in body else None
        run = toolbelt_service.start(
            mode="upload", selectors=selectors
        )
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    return jsonify({"ok": True, "run": run})


@app.route("/api/toolbelt/runs/<run_id>", methods=["GET"])
def toolbelt_run(run_id):
    authorize(Permission.DEPLOY_CERTIFICATE)
    if _toolbelt_unavailable():
        return jsonify({"error": "Toolbelt batch upload is unavailable"}), 503
    run = toolbelt_service.get_run(run_id)
    if run is None:
        return jsonify({"error": "Run not found"}), 404
    return jsonify(run)


@app.route("/api/toolbelt/runs/<run_id>/stop", methods=["POST"])
def toolbelt_stop(run_id):
    authorize(Permission.DEPLOY_CERTIFICATE)
    if _toolbelt_unavailable():
        return jsonify({"error": "Toolbelt batch upload is unavailable"}), 503
    run = toolbelt_service.stop(run_id)
    if run is None:
        return jsonify({"error": "Run not found"}), 404
    return jsonify({"ok": True, "run": run})


@app.route("/api/upload/devices", methods=["GET"])
def list_upload_devices():
    data = load_data()
    devices = data.get("upload_devices", [])
    # Never send password back to frontend
    safe = [{k: v for k, v in d.items() if k != "password"} for d in devices]
    return jsonify(safe)


@app.route("/api/upload/devices", methods=["POST"])
def add_upload_device():
    data = load_data()
    body = request.json
    required = ("name", "host", "device_type")
    if not all(body.get(k, "").strip() for k in required):
        return jsonify({"error": "name, host and device_type are required"}), 400
    device = {
        "id": str(uuid.uuid4()),
        "name": body["name"].strip(),
        "host": body["host"].strip(),
        "device_type": body["device_type"].strip(),
        "username": body.get("username", "").strip(),
        "password": body.get("password", "").strip(),
        "port": int(body.get("port", 80)),
        "https": bool(body.get("https", False)),
        "added": datetime.now(timezone.utc).isoformat(),
    }
    data.setdefault("upload_devices", []).append(device)
    save_data(data)
    safe = {k: v for k, v in device.items() if k != "password"}
    return jsonify({"ok": True, "device": safe})


@app.route("/api/upload/devices/<device_id>", methods=["DELETE"])
def remove_upload_device(device_id):
    data = load_data()
    data["upload_devices"] = [d for d in data.get("upload_devices", []) if d["id"] != device_id]
    save_data(data)
    return jsonify({"ok": True})


@app.route("/api/upload/devices/<device_id>", methods=["PATCH"])
def update_upload_device(device_id):
    data = load_data()
    body = request.json
    for d in data.get("upload_devices", []):
        if d["id"] == device_id:
            for field in ("name", "host", "username", "password", "port", "https", "device_type"):
                if field in body:
                    d[field] = body[field]
            break
    save_data(data)
    return jsonify({"ok": True})


def _extron_push(host, port, use_https, username, password, cert_pem, key_pem):
    """
    Attempt to push a cert+key to an Extron device via its web interface.
    Returns (success: bool, log: list[str])
    """
    import urllib.request
    import urllib.parse
    import urllib.error
    import http.cookiejar

    log = []
    scheme = "https" if use_https else "http"
    base = f"{scheme}://{host}:{port}"

    # Build an opener with cookie jar and no SSL verification
    cj = http.cookiejar.CookieJar()
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    https_handler = urllib.request.HTTPSHandler(context=ctx)
    opener = urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(cj),
        https_handler
    )
    opener.addheaders = [("User-Agent", "Mozilla/5.0")]

    # Step 1: Fetch the login page to discover form fields
    login_candidates = ["/", "/login", "/auth", "/index.html"]
    login_url = None
    for path in login_candidates:
        try:
            resp = opener.open(f"{base}{path}", timeout=5)
            body_bytes = resp.read(4096)
            body_str = body_bytes.decode("utf-8", errors="ignore")
            if "password" in body_str.lower() or "login" in body_str.lower():
                login_url = f"{base}{path}"
                log.append(f"Found login page at {path}")
                break
        except Exception:
            continue

    if not login_url:
        login_url = base + "/"
        log.append("Login page not identified, trying root")

    # Step 2: POST credentials
    login_payloads = [
        {"username": username or "admin", "password": password},
        {"user": username or "admin", "passwd": password},
        {"login": username or "admin", "password": password},
    ]
    login_post_urls = [login_url, base + "/login", base + "/auth", base + "/api/login"]
    authenticated = False
    for post_url in login_post_urls:
        for payload in login_payloads:
            try:
                data_enc = urllib.parse.urlencode(payload).encode()
                req = urllib.request.Request(post_url, data=data_enc,
                                             headers={"Content-Type": "application/x-www-form-urlencoded"})
                resp = opener.open(req, timeout=5)
                status = resp.getcode()
                resp_body = resp.read(1024).decode("utf-8", errors="ignore")
                if status in (200, 302) and "invalid" not in resp_body.lower():
                    log.append(f"Login succeeded at {post_url} (HTTP {status})")
                    authenticated = True
                    break
            except urllib.error.HTTPError as e:
                if e.code == 401:
                    continue
                log.append(f"Login HTTP error {e.code} at {post_url}")
            except Exception as e:
                log.append(f"Login error at {post_url}: {e}")
        if authenticated:
            break

    if not authenticated:
        log.append("WARNING: Could not confirm login — will attempt upload anyway")

    # Step 3: Try known Extron certificate upload endpoints
    cert_bytes = cert_pem.encode() if isinstance(cert_pem, str) else cert_pem
    key_bytes = key_pem.encode() if isinstance(key_pem, str) else key_pem

    boundary = "----CertMonBoundary7a3f9b"
    def make_multipart(cert_field, key_field):
        body = (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="{cert_field}"; filename="cert.pem"\r\n'
            f"Content-Type: application/x-pem-file\r\n\r\n"
        ).encode() + cert_bytes + (
            f"\r\n--{boundary}\r\n"
            f'Content-Disposition: form-data; name="{key_field}"; filename="key.pem"\r\n'
            f"Content-Type: application/x-pem-file\r\n\r\n"
        ).encode() + key_bytes + f"\r\n--{boundary}--\r\n".encode()
        return body

    upload_attempts = [
        ("/api/certificate",        "certificate", "private_key"),
        ("/api/config/certificate", "cert",        "key"),
        ("/Certificate",            "certificate", "key"),
        ("/certificate",            "cert_file",   "key_file"),
        ("/api/security/cert",      "cert",        "key"),
    ]

    for path, cert_field, key_field in upload_attempts:
        try:
            body = make_multipart(cert_field, key_field)
            req = urllib.request.Request(
                base + path,
                data=body,
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}"}
            )
            req.get_method = lambda: "POST"
            resp = opener.open(req, timeout=10)
            status = resp.getcode()
            resp_body = resp.read(512).decode("utf-8", errors="ignore")
            log.append(f"Upload attempt {path}: HTTP {status} — {resp_body[:120]}")
            if status in (200, 201, 204):
                log.append("Certificate upload succeeded!")
                return True, log
        except urllib.error.HTTPError as e:
            log.append(f"Upload {path}: HTTP {e.code}")
            if e.code not in (404, 405):
                # Non-404/405 means the endpoint exists but may need different params
                log.append(f"  → endpoint exists (HTTP {e.code}), may need manual parameter adjustment")
        except Exception as e:
            log.append(f"Upload {path}: {e}")

    log.append("Automatic upload did not succeed — see manual upload instructions")
    return False, log


def _generic_instructions(device):
    device_type = device.get("device_type", "generic")
    host = device.get("host", "")
    port = device.get("port", 443)
    https = device.get("https", False)
    scheme = "https" if https else "http"

    instructions = {
        "extron": (
            f"1. Open Extron Toolbelt on this PC\n"
            f"2. Connect to {host}\n"
            f"3. Go to the Security / Certificate section\n"
            f"4. Upload the .crt file as the Certificate and the .key file as the Private Key\n"
            f"5. Apply and reboot the device if prompted\n\n"
            f"Alternatively: open http://{host} → Security → Certificate"
        ),
        "homeassistant": (
            f"1. Copy cert.pem to your HA config directory (e.g. /config/ssl/fullchain.pem)\n"
            f"2. Copy key.pem to /config/ssl/privkey.pem\n"
            f"3. In configuration.yaml set:\n"
            f"   http:\n"
            f"     ssl_certificate: /config/ssl/fullchain.pem\n"
            f"     ssl_key: /config/ssl/privkey.pem\n"
            f"4. Restart Home Assistant"
        ),
        "synology": (
            f"1. Open DSM → Control Panel → Security → Certificate\n"
            f"2. Click Add → Import certificate\n"
            f"3. Upload cert.pem as Certificate and key.pem as Private Key\n"
            f"4. Set as default if needed"
        ),
        "generic": (
            f"1. Open the device web interface at {scheme}://{host}:{port}\n"
            f"2. Navigate to Security or Certificate settings\n"
            f"3. Upload cert.pem as the certificate and key.pem as the private key\n"
            f"4. Apply / restart as needed"
        ),
    }
    return instructions.get(device_type, instructions["generic"])


@app.route("/api/upload/push", methods=["POST"])
def push_cert():
    authorize(Permission.DEPLOY_CERTIFICATE)
    if deployment_service is None:
        return jsonify({"error": "Secure deployment is unavailable"}), 503
    body = request.get_json(silent=True) or {}
    prohibited = {
        "cert_pem",
        "key_pem",
        "certificate_pem",
        "private_key_pem",
        "combined_pem",
        "private_material",
    }
    if any(field in body for field in prohibited):
        return (
            jsonify(
                {
                    "error": "Private certificate material must not be sent to this API"
                }
            ),
            400,
        )

    device_id = body.get("device_id")
    certificate_id = body.get("certificate_id")
    if not device_id or not certificate_id:
        return jsonify({"error": "certificate_id and device_id are required"}), 400
    unknown = set(body) - {"device_id", "certificate_id"}
    if unknown:
        return jsonify({"error": f"Unsupported fields: {sorted(unknown)}"}), 400

    data = load_data()
    device = next((d for d in data.get("upload_devices", []) if d["id"] == device_id), None)
    if not device:
        return jsonify({"error": "Device not found"}), 404
    try:
        result = deployment_service.deploy_certificate(device, certificate_id)
    except (KeyError, ValueError) as error:
        return jsonify({"error": str(error)}), 400

    payload = {
        "ok": result.ok,
        "log": list(result.log),
        "instructions": result.instructions,
        "public_artifacts": result.public_artifacts,
        "job": _safe_job(result.job) if result.job is not None else None,
    }
    if result.verification is not None:
        payload["verification"] = {
            "status": result.verification.status,
            "expected_fingerprint": result.verification.expected_fingerprint,
            "observed_fingerprint": result.verification.observed_fingerprint,
        }
    return jsonify(payload)


@app.route("/api/upload/test", methods=["POST"])
def test_device_connection():
    body = request.json
    host = body.get("host", "").strip()
    port = int(body.get("port", 80))
    use_https = bool(body.get("https", False))

    if not host:
        return jsonify({"error": "host is required"}), 400

    import urllib.request
    import urllib.error
    scheme = "https" if use_https else "http"
    url = f"{scheme}://{host}:{port}/"

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    opener = urllib.request.build_opener(urllib.request.HTTPSHandler(context=ctx))
    opener.addheaders = [("User-Agent", "Mozilla/5.0")]

    try:
        resp = opener.open(url, timeout=5)
        return jsonify({"ok": True, "status": resp.getcode(), "message": f"Reachable (HTTP {resp.getcode()})"})
    except urllib.error.HTTPError as e:
        # 401/403 means the device is reachable but needs auth — that's fine
        if e.code in (401, 403):
            return jsonify({"ok": True, "status": e.code, "message": f"Reachable — requires authentication (HTTP {e.code})"})
        return jsonify({"ok": False, "status": e.code, "message": f"HTTP {e.code}"})
    except Exception as e:
        return jsonify({"ok": False, "status": 0, "message": str(e)})


if __name__ == "__main__":
    os.makedirs(data_dir(), exist_ok=True)
    port = int(os.environ.get("PORT", 5000))
    print(f"CertMon running at http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
